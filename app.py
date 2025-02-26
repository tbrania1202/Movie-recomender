#!/usr/bin/python
# -*- coding: utf-8 -*-
import string
import numpy as np
import random
from tkinter import *
from tkinter import ttk
from tkinter import filedialog as fd
from tkinter import messagebox
import tkinter.font as font
import main
from tkinter import simpledialog
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from timeit import default_timer as timer
import openpyxl
from itertools import islice

global alternatives_data2
global class_data2
global ranking2
global alternatives_data2_columns
global class_data2_columns
alternatives_data2 = {}
class_data2 = {}
ranking2 = {}
alternatives_data2_columns = {}
class_data2_columns = {}

def set_default():
    global alternatives_data2
    global class_data2
    global ranking2
    global alternatives_data2_columns
    global class_data2_columns
    alternatives_data2 = {}
    class_data2 = {}
    ranking2 = {}
    alternatives_data2_columns = {}
    class_data2_columns = {}
    messagebox.showinfo('Alert', 'Dane zostały wyczyszczone')

def load_data2():
    global alternatives_data2
    file_path = fd.askopenfilename()

    if file_path:
        wb = openpyxl.load_workbook(file_path)

        sheet1 = wb.worksheets[0]
        alternatives_data2 = {}
        for row in sheet1.iter_rows(min_row=2, values_only=True):
            index, name = row[0], row[1]
            values = row[2:]
            alternatives_data2[index] = {'name': name, 'values': values}

        for row in sheet1.iter_rows(min_row=0, max_row=1, values_only=True):
            values = row[0:]
            alternatives_data2_columns[1] = {'values': values}

        wb.close()
        print(alternatives_data2)
        print(alternatives_data2_columns)
        messagebox.showinfo('Alert', 'Dane pobrane pomyślnie!')

def load_criteria2():
    global class_data2
    file_path = fd.askopenfilename()

    if file_path:
        wb = openpyxl.load_workbook(file_path)

        sheet2 = wb.worksheets[0]
        class_data2 = {}
        for row in sheet2.iter_rows(min_row=2, values_only=True):
            class_index = row[0]
            values = row[1:]
            class_data2[class_index] = {'values': values}

        for row in sheet2.iter_rows(min_row=0, max_row=1, values_only=True):
            values = row[0:]
            class_data2_columns[1] = {'values': values}

        wb.close()
        print(class_data2)
        print(class_data2_columns)
        messagebox.showinfo('Alert', 'Klasy pobrane pomyślnie!')

def create_ranking2():
    global si
    global best_rank
    global data_values # niezdominowane
    global ranking2

    print(alternatives_data2)
    print(class_data2)
    print()

    if (alternatives_data2 and class_data2):
        if choose_method2.get() == "TOPSIS":
            data_values = [value['values'] for value in alternatives_data2.values()]
            # data_values, _ = main.owd1(data_values)
            matrix2 = np.array(data_values, dtype="float")
            print(matrix2,"\n")

            weights2 = [1] * len(alternatives_data2[1]['values'])
            weights2 = np.array(weights2, dtype="float")
            print(weights2,"\n")

            data_criteria = [value['values'] for value in class_data2.values()][1]
            criteria2 = np.array(data_criteria, dtype="float")
            print(criteria2,"\n")

            si, best_rank = main.topsis(matrix2, weights2, criteria2, 'euclides')
            for i in range(len(si)):
                ranking2[i+1] = {'name': alternatives_data2[i+1]['name'], 'si':si[i], 'values': alternatives_data2[i+1]['values']}
            print(ranking2)
            ranking2 = dict(sorted(ranking2.items(), key=lambda item: item[1]['si'], reverse=True))
            print("\nRanking: ")
            print(ranking2)

        elif choose_method2.get() == "RSM": 
            data_values = [value['values'] for value in alternatives_data2.values()]
            # data_values, _ = main.owd1(data_values)
            matrix2 = np.array(data_values, dtype="float")
            print(matrix2,"\n")

            weights2 = [1] * len(alternatives_data2[1]['values'])
            weights2 = np.array(weights2, dtype="float")
            print(weights2,"\n")

            data_points2 = [value['values'] for value in class_data2.values()][1]
            points2 = np.array(data_points2, dtype="float")
            print(points2,"\n")

            data_status_quo2 = [value['values'] for value in class_data2.values()][2]
            status_quo2 = np.array(data_status_quo2, dtype="float")
            print(status_quo2,"\n")

            si, best_rank = main.rsm(matrix2, weights2, points2, status_quo2)
            for i in range(len(si)):
                ranking2[i+1] = {'name': alternatives_data2[i+1]['name'], 'si':si[i], 'values': alternatives_data2[i+1]['values']}
            print(ranking2)
            ranking2 = dict(sorted(ranking2.items(), key=lambda item: item[1]['si'], reverse=True))
            print("\nRanking: ")
            print(ranking2)

        elif choose_method2.get() == "UTA STAR":
            data_values = [value['values'] for value in alternatives_data2.values()]
            # data_values, _ = main.owd1(data_values)
            matrix2 = np.array(data_values, dtype="float")
            print(matrix2,"\n")

            weights2 = [1] * len(alternatives_data2[1]['values'])
            weights2 = np.array(weights2, dtype="float")
            print(weights2,"\n")

            data_criteria = [value['values'] for value in class_data2.values()][1]
            criteria2 = np.array(data_criteria, dtype="float")
            print(criteria2,"\n")

            si, best_rank = main.uta_star(matrix2, weights2, criteria2)
            for i in range(len(si)):
                ranking2[i+1] = {'name': alternatives_data2[i+1]['name'], 'si':si[i], 'values': alternatives_data2[i+1]['values']}
            print(ranking2)
            ranking2 = dict(sorted(ranking2.items(), key=lambda item: item[1]['si'], reverse=True))
            print("\nRanking: ")
            print(ranking2)
    else:
        messagebox.showinfo('Alert', 'Nie wprowadzono alternatyw i/lub klas')    

def show_alternatives2():
    if (alternatives_data2):
        database_window = Toplevel()
        database_window.title("Alternatywy z kryteriami")

        database_frame = Frame(database_window)
        database_frame.pack(padx=20, pady=20)

        label = Label(database_frame, text="Alternatywy z kryteriami:")
        label.pack()

        tree = ttk.Treeview(database_frame)

        columns = list(alternatives_data2_columns[next(iter(alternatives_data2_columns))]['values'])
        tree["columns"] = tuple(columns[1:])

        tree.heading("#0", text=columns[0])
        for col in columns[1:]:
            tree.heading(col, text=col)

        for key, value in alternatives_data2.items():
            values = list(value['values'])
            name = value.get('name', '')
            tree.insert("", "end", iid=key, text=key, values=[name]+values)

        tree.pack()
    else:
        messagebox.showinfo('Alert', 'Nie wprowadzono alternatyw')


def show_classes2():
    if (class_data2):
        database_window = Toplevel()
        database_window.title("Klasy")

        database_frame = Frame(database_window)
        database_frame.pack(padx=20, pady=20)

        label = Label(database_frame, text="Klasy:")
        label.pack()

        tree = ttk.Treeview(database_frame)

        columns = list(class_data2_columns[next(iter(class_data2_columns))]['values'])
        tree["columns"] = tuple(columns[1:])

        tree.heading("#0", text=columns[0])
        for col in columns[1:]:
            tree.heading(col, text=col)

        for key, value in class_data2.items():
            values = list(value['values'])
            tree.insert("", "end", iid=key, text=key, values=values)

        tree.pack()
    else:
        messagebox.showinfo('Alert', 'Nie wprowadzono klas')

def show_ranking2():
    database_window = Toplevel()
    database_window.title("Ranking")

    database_frame = Frame(database_window)
    database_frame.pack(padx=20, pady=20)

    label = Label(database_frame, text="Ranking:")
    label.pack()

    tree = ttk.Treeview(database_frame)

    columns = ["Indeks z bazy"] + ["Tytuł"] + ["Współczynnik"]
    tree["columns"] = tuple(columns)

    for col in columns:
        tree.heading(col, text=col)

    tree.column("#0", width=0, stretch=NO)
    tree.column("#1", anchor=CENTER)

    for key, value in ranking2.items():
        values = [key, value['name'], value['si']]
        tree.insert("", "end", values=values)

    tree.pack()

def display2():
    if len(alternatives_data2):
        ranking_to_display2 = dict(islice(ranking2.items(), 10))

        if len(alternatives_data2[1]['values']) == 2:
            top = Toplevel()
            fig = plt.figure()
            ax = fig.add_subplot(111)
            x = [value['values'][0] for value in ranking_to_display2.values()]
            y = [value['values'][1] for value in ranking_to_display2.values()]
            ax.scatter(x, y, label='Najlepsze 10 punktów', marker='*')
            ax.set_xlabel("Kryterium 1")
            ax.set_ylabel("Kryterium 2")
            ax.legend()
            canvas = FigureCanvasTkAgg(fig, master=top)
            canvas.get_tk_widget().pack()
            canvas.draw()
            
        elif len(alternatives_data2[1]['values']) == 3:
            top = Toplevel()
            fig = plt.figure()
            ax = fig.add_subplot(111, projection='3d')
            x = [value['values'][0] for value in ranking_to_display2.values()]
            y = [value['values'][1] for value in ranking_to_display2.values()]
            z = [value['values'][2] for value in ranking_to_display2.values()]
            ax.scatter(x, y, z, label='Najlepsze 10 punktów', marker='*')
            ax.set_xlabel("Kryterium 1")
            ax.set_ylabel("Kryterium 2")
            ax.set_zlabel("Kryterium 3")
            ax.legend()
            canvas = FigureCanvasTkAgg(fig, master=top)
            canvas.get_tk_widget().pack()
            canvas.draw()

        elif len(alternatives_data2[1]['values']) == 4:
            top = Toplevel()
            fig = plt.figure()
            ax = fig.add_subplot(111, projection='3d')
            x = [value['values'][0] for value in ranking_to_display2.values()]
            y = [value['values'][1] for value in ranking_to_display2.values()]
            z = [value['values'][2] for value in ranking_to_display2.values()]
            leg = [value['values'][3] for value in ranking_to_display2.values()]
            colormap = plt.cm.turbo
            norm = plt.Normalize(min(leg), max(leg))
            sc = ax.scatter(x, y, z, c=leg, label='Najlepsze 10 punktów', cmap=colormap, norm=norm, marker='*')
            cbar = plt.colorbar(sc)
            ax.set_xlabel("Kryterium 1")
            ax.set_ylabel("Kryterium 2")
            ax.set_zlabel("Kryterium 3")
            ax.legend()
            cbar.set_label('Kryterium 4')
            canvas = FigureCanvasTkAgg(fig, master=top)
            canvas.get_tk_widget().pack()
            canvas.draw()

        else:
            messagebox.showinfo('Alert', 'Wykresy możliwe tylko dla wymiarów 2D, 3D oraz 4D')
    else:
        messagebox.showinfo('Alert', 'Brak danych do wizualizacji')


root = Tk()
root.title('Optymalizacja wielokryterialna')

myFont = font.Font(family='Helvetica', size=15, weight='bold')

default = Button(root, text='Wyczyść zapis', command=set_default, font=myFont, fg='black', bg='gray')
default.grid(row=5, column=2, padx=5, pady=5)

end_program = Button(root, text='Zakończ', command=lambda: root.destroy(), font=myFont, fg='black', bg='gray')
end_program.grid(row=5, column=3, padx=5, pady=5)

load_data_button2 = Button(root, text='Wczytaj dane z pliku', command=load_data2, font=myFont, fg='black', bg='light gray')
load_data_button2.grid(row=1, column=0, padx=5, pady=5)

load_data_button2 = Button(root, text='Wczytaj kryteria z pliku', command=load_criteria2, font=myFont, fg='black', bg='light gray')
load_data_button2.grid(row=2, column=0, padx=5, pady=5)

choose_method2 = StringVar()
choose_method2.set('Ustaw metodę:')
lst2 = ('TOPSIS', 'RSM', 'UTA STAR')
options2 = ttk.OptionMenu(root, choose_method2, '', *lst2)
options2.grid(row=2, column=1, padx=5, pady=5)

create_ranking_button2 = Button(root, text='Stwórz ranking', command=create_ranking2, font=myFont, fg='black', bg='light gray')
create_ranking_button2.grid(row=3, column=1, padx=5, pady=5)

show_alternatives_button2 = Button(root, text='Pokaż alternatywy', command=show_alternatives2, font=myFont, fg='black', bg='light gray')
show_alternatives_button2.grid(row=3, column=0, padx=5, pady=5)

show_classes_button2 = Button(root, text='Pokaż klasy', command=show_classes2, font=myFont, fg='black', bg='light gray')
show_classes_button2.grid(row=4, column=0, padx=5, pady=5)

show_ranking_button2 = Button(root, text='Pokaż ranking', command=show_ranking2, font=myFont, fg='black', bg='light gray')
show_ranking_button2.grid(row=4, column=1, padx=5, pady=5)

display_button2 = Button(root, text='Wizualizacja', command=display2, font=myFont, fg='black', bg='light gray')
display_button2.grid(row=4, column=2, padx=5, pady=5)

root.mainloop()
